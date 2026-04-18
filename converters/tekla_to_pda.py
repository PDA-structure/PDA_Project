"""
Convert a Tekla Structural Designer (TSD) Excel export to PDA canonical JSON.

Usage:
    python converters/tekla_to_pda.py path/to/tsd_export.xlsx [--output path/to/output.json]

The COLUMN_MAP below must be adjusted to match your TSD version's column headers
and sheet names. TSD column headers vary by version and locale.

The output JSON conforms to PDA schema_version "1.0" for the frame2d solver.
Supports and loads are NOT imported from TSD (TSD exports these separately in
different sheets and the mapping is version-specific) — add them in the PDA
browser UI after loading the generated JSON.
"""

import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    print(
        "openpyxl is required. Install it: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


# ----- USER-CONFIGURABLE: adjust to match your TSD export -----
# Sheet names and column headers vary by TSD version/locale.
# Open your .xlsx in Excel, look at the Nodes and Elements sheets, and copy
# the exact header strings below.
COLUMN_MAP = {
    "nodes_sheet":   "Nodes",
    "node_id":       "Node",
    "node_x":        "X (m)",
    "node_y":        "Y (m)",
    "members_sheet": "Elements",
    "member_id":     "Element",
    "member_start":  "Start Node",
    "member_end":    "End Node",
    "member_E":      "E (kN/m2)",
    "member_I":      "I (m4)",
    "member_A":      "A (m2)",
}
# ---------------------------------------------------------------


def read_sheet_as_dicts(ws):
    """Return list of row dicts keyed by header row values, skipping blank rows.

    The first row is treated as the header row. Rows where every cell is None
    are skipped (openpyxl returns these for trailing empty rows).
    """
    headers = [cell.value for cell in ws[1]]
    return [
        {headers[i]: row[i].value for i in range(len(headers))}
        for row in ws.iter_rows(min_row=2)
        if any(cell.value is not None for cell in row)
    ]


def _require_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            "Sheet '{}' not found in workbook. Available sheets: {}. "
            "Adjust COLUMN_MAP at the top of tekla_to_pda.py to match your "
            "TSD export.".format(sheet_name, wb.sheetnames)
        )
    return wb[sheet_name]


def _require_column(row_dict, column_name, row_index, sheet_name):
    if column_name not in row_dict:
        raise KeyError(
            "Column '{}' not found in sheet '{}' (row {}). Columns present: {}. "
            "Adjust COLUMN_MAP to match your TSD export.".format(
                column_name, sheet_name, row_index + 2, list(row_dict.keys())
            )
        )
    return row_dict[column_name]


def _convert_E_to_Pa(value, header_name):
    """Convert a modulus value to Pa based on the column-header unit hint.

    Recognised units (case-insensitive substring match in the header):
      - "kN/m2"  -> multiply by 1_000   (kilopascal to pascal)
      - "MPa"    -> multiply by 1_000_000
      - "GPa"    -> multiply by 1_000_000_000
      - "Pa"     -> pass-through
    Default (unrecognised): pass-through with no scaling.
    """
    if value is None:
        return None
    h = header_name.lower() if header_name else ""
    v = float(value)
    if "kn/m2" in h or "kn/m^2" in h:
        return v * 1_000.0
    if "gpa" in h:
        return v * 1_000_000_000.0
    if "mpa" in h:
        return v * 1_000_000.0
    # Pa or unknown — pass through
    return v


def _collapse_if_uniform(values):
    """Return scalar if all values equal, else list. Empty list -> None."""
    if not values:
        return None
    first = values[0]
    for v in values:
        if v != first:
            return list(values)
    return first


def convert(xlsx_path, column_map=None):
    """Convert a TSD Excel export to the PDA canonical JSON schema (as a dict).

    Parameters
    ----------
    xlsx_path : str
        Path to the TSD .xlsx file.
    column_map : dict, optional
        Override of COLUMN_MAP (use defaults otherwise).

    Returns
    -------
    dict
        Matches PDA schema_version "1.0" for the frame2d solver. Supports and
        loads are empty (user to add in UI).
    """
    cmap = column_map if column_map is not None else COLUMN_MAP

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # ---------- Nodes ----------
    nodes_ws = _require_sheet(wb, cmap["nodes_sheet"])
    node_rows = read_sheet_as_dicts(nodes_ws)

    tsd_id_to_index = {}
    nodes = []
    for i, row in enumerate(node_rows):
        nid = _require_column(row, cmap["node_id"], i, cmap["nodes_sheet"])
        x = _require_column(row, cmap["node_x"], i, cmap["nodes_sheet"])
        y = _require_column(row, cmap["node_y"], i, cmap["nodes_sheet"])
        if nid is None or x is None or y is None:
            continue
        tsd_id_to_index[nid] = len(nodes)  # contiguous 0-based index
        nodes.append([float(x), float(y)])

    n_nodes = len(nodes)

    # ---------- Members ----------
    members_ws = _require_sheet(wb, cmap["members_sheet"])
    member_rows = read_sheet_as_dicts(members_ws)

    members = []
    E_values = []
    I_values = []
    A_values = []
    E_header = cmap.get("member_E", "")

    for i, row in enumerate(member_rows):
        start_id = _require_column(row, cmap["member_start"], i, cmap["members_sheet"])
        end_id = _require_column(row, cmap["member_end"], i, cmap["members_sheet"])
        if start_id is None or end_id is None:
            continue
        if start_id not in tsd_id_to_index:
            raise ValueError(
                "Member row {} references start node ID {} which is not in the "
                "Nodes sheet.".format(i + 2, start_id)
            )
        if end_id not in tsd_id_to_index:
            raise ValueError(
                "Member row {} references end node ID {} which is not in the "
                "Nodes sheet.".format(i + 2, end_id)
            )
        # PDA API uses 1-based node indices
        members.append(
            [tsd_id_to_index[start_id] + 1, tsd_id_to_index[end_id] + 1]
        )

        # Properties (optional)
        E_raw = row.get(cmap["member_E"])
        I_raw = row.get(cmap["member_I"])
        A_raw = row.get(cmap["member_A"])
        if E_raw is not None:
            E_values.append(_convert_E_to_Pa(E_raw, E_header))
        if I_raw is not None:
            I_values.append(float(I_raw))
        if A_raw is not None:
            A_values.append(float(A_raw))

    n_members = len(members)

    # Collapse uniform property lists to scalars. If column missing entirely,
    # fall back to generic steel defaults so the output is still solve-ready.
    E_out = _collapse_if_uniform(E_values) if len(E_values) == n_members else 200e9
    I_out = _collapse_if_uniform(I_values) if len(I_values) == n_members else 1e-4
    A_out = _collapse_if_uniform(A_values) if len(A_values) == n_members else 0.01

    # ---------- Assemble canonical schema ----------
    result = {
        "schema_version": "1.0",
        "solver": "frame2d",
        "nodes": nodes,
        "members": members,
        "ENForces": [[0, 0] for _ in range(n_members)],
        "ENMoments": [[0, 0] for _ in range(n_members)],
        "forceVector": [0] * (n_nodes * 3),  # 3 DOF/node for frame2d
        "E": E_out,
        "I": I_out,
        "A": A_out,
        "bars": [],
        "beamPinLeft": [],
        "beamPinRight": [],
        "restrainedDoF": [],  # TSD supports not imported — user adds in UI
        "pinDoF": [],
        "springDoF": [],
        "springStiffness": [],
        "udl_x": [0] * n_members,
        "canvas": {
            "origin": None,
            "nodes": [],
            "members": [],
            "supports": [],
            "nodeLoads": [],
        },
    }
    return result


def main():
    parser = argparse.ArgumentParser(
        description="Convert a Tekla Structural Designer Excel export to PDA canonical JSON",
    )
    parser.add_argument("xlsx_path", help="Path to TSD .xlsx export file")
    parser.add_argument(
        "--output",
        "-o",
        help="Output JSON file path (default: stdout)",
    )
    args = parser.parse_args()

    result = convert(args.xlsx_path)

    if args.output:
        with open(args.output, "w") as f:
            json.dump(result, f, indent=2)
        print("Written to {}".format(args.output), file=sys.stderr)
    else:
        print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
