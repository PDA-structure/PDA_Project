"""
Tests for the Tekla Structural Designer Excel converter.

Uses hand-crafted openpyxl workbooks as fixtures (no actual TSD export needed).
Covers schema shape, node-ID remapping (Pitfall 3), forceVector length, and
E-unit conversion.
"""

import os
import sys

import openpyxl
import pytest

# Ensure the worktree / project root is on sys.path so `converters` is importable.
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from converters.tekla_to_pda import convert, read_sheet_as_dicts, COLUMN_MAP  # noqa: E402


# ---------- Fixtures ----------

@pytest.fixture
def minimal_tsd_xlsx(tmp_path):
    """Hand-crafted Excel with 3 nodes, 2 members in TSD column format."""
    wb = openpyxl.Workbook()
    ws_nodes = wb.active
    ws_nodes.title = "Nodes"
    ws_nodes.append(["Node", "X (m)", "Y (m)"])
    ws_nodes.append([1, 0.0, 0.0])
    ws_nodes.append([2, 3.0, 0.0])
    ws_nodes.append([3, 3.0, 4.0])

    ws_members = wb.create_sheet("Elements")
    ws_members.append(
        ["Element", "Start Node", "End Node", "E (kN/m2)", "I (m4)", "A (m2)"]
    )
    ws_members.append([1, 1, 2, 200e6, 1e-4, 0.01])
    ws_members.append([2, 2, 3, 200e6, 1e-4, 0.01])

    path = tmp_path / "tsd_export.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def noncontiguous_tsd_xlsx(tmp_path):
    """TSD export with non-contiguous node IDs (gaps from deleted nodes)."""
    wb = openpyxl.Workbook()
    ws_nodes = wb.active
    ws_nodes.title = "Nodes"
    ws_nodes.append(["Node", "X (m)", "Y (m)"])
    ws_nodes.append([5, 0.0, 0.0])     # ID=5, not 1
    ws_nodes.append([12, 3.0, 0.0])    # ID=12, not 2
    ws_nodes.append([27, 3.0, 4.0])    # ID=27, not 3

    ws_members = wb.create_sheet("Elements")
    ws_members.append(
        ["Element", "Start Node", "End Node", "E (kN/m2)", "I (m4)", "A (m2)"]
    )
    ws_members.append([1, 5, 12, 200e6, 1e-4, 0.01])
    ws_members.append([2, 12, 27, 200e6, 1e-4, 0.01])

    path = tmp_path / "tsd_noncontiguous.xlsx"
    wb.save(path)
    return str(path)


# ---------- Tests ----------

def test_convert_produces_valid_schema(minimal_tsd_xlsx):
    result = convert(minimal_tsd_xlsx)
    assert result["schema_version"] == "1.0"
    assert result["solver"] == "frame2d"
    assert len(result["nodes"]) == 3
    assert len(result["members"]) == 2
    assert "canvas" in result


def test_convert_node_coordinates(minimal_tsd_xlsx):
    result = convert(minimal_tsd_xlsx)
    assert result["nodes"][0] == [0.0, 0.0]
    assert result["nodes"][1] == [3.0, 0.0]
    assert result["nodes"][2] == [3.0, 4.0]


def test_convert_members_are_1_based(minimal_tsd_xlsx):
    result = convert(minimal_tsd_xlsx)
    assert result["members"][0] == [1, 2]
    assert result["members"][1] == [2, 3]


def test_convert_noncontiguous_ids_remapped(noncontiguous_tsd_xlsx):
    result = convert(noncontiguous_tsd_xlsx)
    # TSD IDs 5 -> 1, 12 -> 2, 27 -> 3
    assert result["members"][0] == [1, 2]
    assert result["members"][1] == [2, 3]
    assert len(result["nodes"]) == 3


def test_convert_force_vector_length(minimal_tsd_xlsx):
    result = convert(minimal_tsd_xlsx)
    # frame2d: 3 DOF per node
    assert len(result["forceVector"]) == 3 * 3


def test_convert_e_unit_conversion(minimal_tsd_xlsx):
    """E column header 'E (kN/m2)' with value 200e6 should convert to 200e9 Pa."""
    result = convert(minimal_tsd_xlsx)
    # All members share E, so convert returns a scalar
    E = result["E"]
    assert E == pytest.approx(200e9)


def test_read_sheet_as_dicts(minimal_tsd_xlsx):
    wb = openpyxl.load_workbook(minimal_tsd_xlsx, data_only=True)
    rows = read_sheet_as_dicts(wb["Nodes"])
    assert len(rows) == 3
    for r in rows:
        assert "Node" in r
        assert "X (m)" in r
        assert "Y (m)" in r
    assert rows[0]["Node"] == 1
    assert rows[1]["X (m)"] == 3.0


def test_convert_default_fields_are_present(minimal_tsd_xlsx):
    """All canonical-schema fields required by Frame2DRequest must be present."""
    result = convert(minimal_tsd_xlsx)
    required = [
        "schema_version", "solver", "nodes", "members",
        "ENForces", "ENMoments", "forceVector",
        "E", "I", "A",
        "bars", "beamPinLeft", "beamPinRight",
        "restrainedDoF", "pinDoF", "springDoF", "springStiffness",
        "udl_x", "canvas",
    ]
    for field in required:
        assert field in result, "missing field: {}".format(field)

    # ENForces / ENMoments / udl_x are per-member arrays of the right length
    n_members = len(result["members"])
    assert len(result["ENForces"]) == n_members
    assert len(result["ENMoments"]) == n_members
    assert len(result["udl_x"]) == n_members
    for ef in result["ENForces"]:
        assert ef == [0, 0]
    for em in result["ENMoments"]:
        assert em == [0, 0]


def test_convert_canvas_is_empty_for_tsd(minimal_tsd_xlsx):
    """TSD conversion produces a geometry-only payload — canvas is empty."""
    result = convert(minimal_tsd_xlsx)
    canvas = result["canvas"]
    assert canvas["origin"] is None
    assert canvas["nodes"] == []
    assert canvas["members"] == []
    assert canvas["supports"] == []
    assert canvas["nodeLoads"] == []
